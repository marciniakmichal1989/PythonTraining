import openpyxl
from openpyxl.chart import BarChart, Reference

def prcess_workbook(file_name):
    #pojedynczy plik
    #wb = openpyxl.load_workbook("danePogodowe.xlsx")
    wb = openpyxl.load_workbook(file_name)
    sheet = wb["Arkusz1"]

    #cell = sheet['a1']
    #to samo co nizej
    #cell = sheet.cell(1,1)
    #print(cell.value)
    #print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        #print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chard = BarChart()
    chard.add_data(values)
    sheet.add_chart(chard,'f2')

    #towrzenie nowego pliku
    #wb.save('danePogodowe2.xlsx')
    #napisywanie starego
    wb.save(file_name)

prcess_workbook("danePogodowe.xlsx")
